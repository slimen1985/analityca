import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import PatternFill


# Читаем данные из файла
df = pd.read_excel("analitics_all.xlsx", sheet_name="Заявки и график")

# Получаем список уникальных покупателей
buyers = df["покупатель"].unique()

# Создаем новый файл для каждого покупателя
for buyer in buyers:
    # Отбираем строки, относящиеся к данному покупателю
    buyer_data = df[df["покупатель"] == buyer]
    # Создаем новый файл и записываем данные
    new_file_name = f"{buyer}.xlsx"
    with pd.ExcelWriter(new_file_name) as writer:
        buyer_data.to_excel(writer, index=False)
        # Задаем ширину колонок в соответствии с содержимым
        worksheet = writer.sheets["Sheet1"]
        for idx, col in enumerate(worksheet.columns):
            max_length = 0
            column = col[0].column_letter  # Получаем букву столбца
            for cell in col:
                try:  # Обрабатываем ошибки
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2)
            worksheet.column_dimensions[column].width = adjusted_width

            # Устанавливаем цвет ячейки в соответствии с содержимым колонки "Заявлено раз"
            if column == "J":
                for cell in col:
                    try:
                        value = int(cell.value)
                        if value >= 4:
                            fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
                        elif value == 3:
                            fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
                        elif value == 2:
                            fill = PatternFill(start_color="FFFF99", end_color="FFFF99", fill_type="solid")
                        elif value == 1:
                            fill = PatternFill(start_color="CCFFCC", end_color="CCFFCC", fill_type="solid")
                        else:
                            fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
                        cell.fill = fill
                    except:
                        pass

        # Добавляем автофильтр
        worksheet.auto_filter.ref = worksheet.dimensions

    print(f"Файл {new_file_name} успешно создан")
